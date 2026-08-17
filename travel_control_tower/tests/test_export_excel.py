from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory
import unittest

from openpyxl import load_workbook

from travel_control_tower.exporters.excel_export import export_plan_to_excel
from travel_control_tower.planner_core.models import TripRequest
from travel_control_tower.planner_core.pipeline import build_plan_stub


class ExcelExportTests(unittest.TestCase):
    def test_export_plan_to_excel_creates_expected_sheets(self) -> None:
        request = TripRequest(
            scenario_id="japan_osaka_weekend",
            departure_city="上海",
            destination="大阪",
            start_date="2026-06-09",
            end_date="2026-06-11",
            traveler_count=1,
            budget_per_person=3800,
        )
        plan = build_plan_stub(request).to_dict()

        with TemporaryDirectory() as tmp_dir:
            output_path = Path(tmp_dir) / "plan.xlsx"
            export_plan_to_excel(plan, output_path)

            self.assertTrue(output_path.exists())

            workbook = load_workbook(output_path)
            self.assertEqual(
                workbook.sheetnames,
                ["总览", "当前选择", "预算", "预定事项", "携程飞猪对比", "行程明细", "行程甘特图", "路线甘特图"],
            )
            self.assertEqual(workbook["总览"]["A1"].value, plan["overview"]["title"])
            self.assertEqual(workbook["当前选择"]["A1"].value, "当前选择")
            self.assertEqual(workbook["预算"]["A1"].value, "预算")
            self.assertEqual(workbook["预定事项"]["A1"].value, "预定事项")
            self.assertEqual(workbook["预定事项"]["E3"].value, "为什么现在订")
            self.assertEqual(workbook["预定事项"]["F3"].value, "拖晚的风险")
            self.assertEqual(workbook["携程飞猪对比"]["F4"].value, "携程入口")
            self.assertEqual(workbook["携程飞猪对比"]["G4"].value, "飞猪入口")
            self.assertEqual(workbook["行程明细"]["K3"].value, "交通方式")
            self.assertEqual(workbook["行程明细"]["L3"].value, "起点")
            self.assertEqual(workbook["行程明细"]["M3"].value, "终点")
            self.assertEqual(workbook["行程明细"]["N3"].value, "距离(km)")
            self.assertEqual(workbook["行程明细"]["O3"].value, "路线来源")
            self.assertEqual(workbook["行程明细"]["P3"].value, "路线说明")


if __name__ == "__main__":
    unittest.main()
