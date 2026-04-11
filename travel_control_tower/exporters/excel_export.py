from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


CATEGORY_COLORS = {
    "交通": "DCEBFA",
    "游玩": "E7F5E8",
    "餐食": "FCE8CC",
    "住宿": "EDE2FF",
    "缓冲": "E5E7EB",
}


def export_plan_to_excel(plan: dict, output_path: Path) -> Path:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    overview = workbook.active
    overview.title = "总览"

    _write_overview(overview, plan)
    _write_selected(workbook.create_sheet("当前选择"), plan)
    _write_budget(workbook.create_sheet("预算"), plan)
    _write_booking(workbook.create_sheet("预定事项"), plan)
    _write_itinerary(workbook.create_sheet("行程明细"), plan)
    _write_gantt(workbook.create_sheet("行程甘特图"), plan)
    _write_route_gantt(workbook.create_sheet("路线甘特图"), plan)

    workbook.save(output_path)
    return output_path


def _title_cell(sheet, row: int, title: str) -> None:
    sheet.cell(row=row, column=1, value=title)
    sheet.cell(row=row, column=1).font = Font(size=16, bold=True)


def _write_overview(sheet, plan: dict) -> None:
    overview = plan.get("overview", {})
    input_snapshot = plan.get("input_snapshot", {})
    statuses = plan.get("provider_statuses", [])
    assumptions = plan.get("assumptions", [])

    _title_cell(sheet, 1, overview.get("title", "旅行方案"))
    sheet.cell(row=2, column=1, value="方案摘要")
    sheet.cell(row=2, column=2, value=overview.get("summary", ""))

    row = 4
    sheet.cell(row=row, column=1, value="本次输入已识别为")
    sheet.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    for key, value in input_snapshot.items():
        sheet.cell(row=row, column=1, value=key)
        sheet.cell(row=row, column=2, value=_stringify(value))
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="数据状态")
    sheet.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    for item in statuses:
        sheet.cell(row=row, column=1, value=item.get("name", ""))
        sheet.cell(row=row, column=2, value=item.get("status", ""))
        sheet.cell(row=row, column=3, value=item.get("details", ""))
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="当前假设")
    sheet.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    for item in assumptions:
        sheet.cell(row=row, column=1, value="•")
        sheet.cell(row=row, column=2, value=item)
        row += 1

    _set_widths(sheet, {1: 20, 2: 48, 3: 72})
    _wrap_all(sheet)


def _write_selected(sheet, plan: dict) -> None:
    _title_cell(sheet, 1, "当前选择")

    hotel = plan.get("selected_hotel") or {}
    transport = plan.get("selected_transport") or {}

    row = 3
    sheet.cell(row=row, column=1, value="主酒店")
    sheet.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    for label, value in [
        ("名称", hotel.get("name", "")),
        ("价格", hotel.get("nightly_price", "")),
        ("区域", hotel.get("area", "")),
        ("说明", hotel.get("notes", "")),
        ("链接", hotel.get("booking_url", "")),
    ]:
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=_stringify(value))
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="主交通")
    sheet.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    for label, value in [
        ("名称", transport.get("label", "")),
        ("分类", transport.get("category", "")),
        ("总价", transport.get("total_price", "")),
        ("出发", transport.get("depart_at", "")),
        ("到达", transport.get("arrive_at", "")),
        ("链接", transport.get("booking_url", "")),
    ]:
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=_stringify(value))
        row += 1

    row += 2
    sheet.cell(row=row, column=1, value="酒店候选")
    sheet.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    for col, header in enumerate(["名称", "价格", "区域", "说明", "链接"], start=1):
        sheet.cell(row=row, column=col, value=header)
        sheet.cell(row=row, column=col).font = Font(bold=True)
    row += 1
    for item in plan.get("hotel_candidates", []):
        sheet.cell(row=row, column=1, value=item.get("name", ""))
        sheet.cell(row=row, column=2, value=item.get("nightly_price", ""))
        sheet.cell(row=row, column=3, value=item.get("area", ""))
        sheet.cell(row=row, column=4, value=item.get("notes", ""))
        sheet.cell(row=row, column=5, value=item.get("booking_url", ""))
        row += 1

    row += 1
    sheet.cell(row=row, column=1, value="交通候选")
    sheet.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    for col, header in enumerate(["名称", "分类", "总价", "出发", "到达", "链接"], start=1):
        sheet.cell(row=row, column=col, value=header)
        sheet.cell(row=row, column=col).font = Font(bold=True)
    row += 1
    for item in plan.get("transport_candidates", []):
        sheet.cell(row=row, column=1, value=item.get("label", ""))
        sheet.cell(row=row, column=2, value=item.get("category", ""))
        sheet.cell(row=row, column=3, value=item.get("total_price", ""))
        sheet.cell(row=row, column=4, value=item.get("depart_at", ""))
        sheet.cell(row=row, column=5, value=item.get("arrive_at", ""))
        sheet.cell(row=row, column=6, value=item.get("booking_url", ""))
        row += 1

    _set_widths(sheet, {1: 18, 2: 18, 3: 18, 4: 42, 5: 58, 6: 58})
    _wrap_all(sheet)


def _write_budget(sheet, plan: dict) -> None:
    budget = plan.get("budget", {})
    _title_cell(sheet, 1, "预算")
    summary_rows = [
        ("总预算", budget.get("fixed_cost_total", 0)),
        ("人均预算", budget.get("per_person_cost", 0)),
        ("可升级项", budget.get("optional_upgrade_total", 0)),
    ]
    row = 3
    for label, value in summary_rows:
        sheet.cell(row=row, column=1, value=label)
        sheet.cell(row=row, column=2, value=value)
        row += 1

    row += 1
    headers = ["分类", "总额", "人均", "说明"]
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=row, column=col, value=header)
        sheet.cell(row=row, column=col).font = Font(bold=True)
    row += 1

    for item in budget.get("breakdown", []):
        sheet.cell(row=row, column=1, value=item.get("category", ""))
        sheet.cell(row=row, column=2, value=item.get("total", 0))
        sheet.cell(row=row, column=3, value=item.get("per_person", 0))
        sheet.cell(row=row, column=4, value=item.get("notes", ""))
        row += 1

    _set_widths(sheet, {1: 18, 2: 16, 3: 16, 4: 64})
    _wrap_all(sheet)


def _write_booking(sheet, plan: dict) -> None:
    _title_cell(sheet, 1, "预定事项")
    headers = ["事项", "类型", "优先级", "建议时间", "为什么现在订", "拖晚的风险", "说明", "链接"]
    row = 3
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=row, column=col, value=header)
        sheet.cell(row=row, column=col).font = Font(bold=True)
    row += 1

    for item in plan.get("booking_items", []):
        sheet.cell(row=row, column=1, value=item.get("name", ""))
        sheet.cell(row=row, column=2, value=item.get("category", ""))
        sheet.cell(row=row, column=3, value=item.get("priority", ""))
        sheet.cell(row=row, column=4, value=item.get("timing", ""))
        sheet.cell(row=row, column=5, value=item.get("why_now", ""))
        sheet.cell(row=row, column=6, value=item.get("risk_if_wait", ""))
        sheet.cell(row=row, column=7, value=item.get("notes", ""))
        sheet.cell(row=row, column=8, value=item.get("url", ""))
        row += 1

    _set_widths(sheet, {1: 32, 2: 14, 3: 12, 4: 18, 5: 28, 6: 28, 7: 52, 8: 64})
    _wrap_all(sheet)


def _write_itinerary(sheet, plan: dict) -> None:
    _title_cell(sheet, 1, "行程明细")
    headers = [
        "天数",
        "日期",
        "主题",
        "当天预计花费",
        "开始",
        "结束",
        "时长(分钟)",
        "分类",
        "项目",
        "说明",
        "交通方式",
        "起点",
        "终点",
        "距离(km)",
        "路线来源",
        "路线说明",
    ]
    row = 3
    for col, header in enumerate(headers, start=1):
        sheet.cell(row=row, column=col, value=header)
        sheet.cell(row=row, column=col).font = Font(bold=True)
    row += 1

    for day in plan.get("daily_plan", []):
        first_row = row
        for item in day.get("items", []):
            sheet.cell(row=row, column=1, value=day.get("day_index", ""))
            sheet.cell(row=row, column=2, value=day.get("date", ""))
            sheet.cell(row=row, column=3, value=day.get("theme", ""))
            sheet.cell(row=row, column=4, value=day.get("estimated_cost_total", 0))
            sheet.cell(row=row, column=5, value=item.get("start_time", ""))
            sheet.cell(row=row, column=6, value=item.get("end_time", ""))
            sheet.cell(row=row, column=7, value=item.get("duration_minutes", 0))
            sheet.cell(row=row, column=8, value=item.get("category", ""))
            sheet.cell(row=row, column=9, value=item.get("label", ""))
            sheet.cell(row=row, column=10, value=item.get("notes", ""))
            sheet.cell(row=row, column=11, value=item.get("route_mode_label") or item.get("route_mode", ""))
            sheet.cell(row=row, column=12, value=item.get("route_origin", ""))
            sheet.cell(row=row, column=13, value=item.get("route_destination", ""))
            sheet.cell(row=row, column=14, value=_route_distance_text(item.get("route_distance_km", "")))
            sheet.cell(row=row, column=15, value=_display_route_provider(item.get("route_provider", "")))
            sheet.cell(row=row, column=16, value=item.get("route_summary", ""))
            row += 1

        sheet.cell(row=first_row, column=18, value="为什么这样排")
        sheet.cell(row=first_row, column=19, value=day.get("why_this_day", ""))
        sheet.cell(row=first_row + 1, column=18, value="交通策略")
        sheet.cell(row=first_row + 1, column=19, value=day.get("transport_strategy", ""))
        sheet.cell(row=first_row + 2, column=18, value="餐食策略")
        sheet.cell(row=first_row + 2, column=19, value=day.get("meal_strategy", ""))
        sheet.cell(row=first_row + 3, column=18, value="进度快怎么办")
        sheet.cell(row=first_row + 3, column=19, value=day.get("fallback_if_fast", ""))
        sheet.cell(row=first_row + 4, column=18, value="太累怎么办")
        sheet.cell(row=first_row + 4, column=19, value=day.get("fallback_if_tired", ""))
        sheet.cell(row=first_row + 5, column=18, value="当天花费说明")
        sheet.cell(row=first_row + 5, column=19, value=day.get("estimated_cost_notes", ""))
        row += 1

    _set_widths(
        sheet,
        {
            1: 8,
            2: 14,
            3: 20,
            4: 14,
            5: 10,
            6: 10,
            7: 12,
            8: 12,
            9: 24,
            10: 48,
            11: 14,
            12: 18,
            13: 18,
            14: 12,
            15: 14,
            16: 42,
            18: 14,
            19: 54,
        },
    )
    _wrap_all(sheet)


def _write_gantt(sheet, plan: dict) -> None:
    _title_cell(sheet, 1, "甘特图")
    start_minutes = 8 * 60
    end_minutes = 22 * 60
    slot_minutes = 30

    sheet.cell(row=3, column=1, value="日期")
    sheet.cell(row=3, column=2, value="主题")
    sheet.cell(row=3, column=1).font = Font(bold=True)
    sheet.cell(row=3, column=2).font = Font(bold=True)

    slot_count = int((end_minutes - start_minutes) / slot_minutes)
    for idx in range(slot_count):
        minute = start_minutes + idx * slot_minutes
        hour = minute // 60
        mins = minute % 60
        col = 3 + idx
        sheet.cell(row=3, column=col, value=f"{hour:02d}:{mins:02d}")
        sheet.cell(row=3, column=col).font = Font(bold=True, size=9)
        sheet.column_dimensions[get_column_letter(col)].width = 5

    row = 4
    for day in plan.get("daily_plan", []):
        sheet.cell(row=row, column=1, value=day.get("date", ""))
        sheet.cell(row=row, column=2, value=day.get("theme", ""))

        for item in day.get("items", []):
            start_time = item.get("start_time", "")
            end_time = item.get("end_time", "")
            if not start_time or not end_time:
                continue
            start_slot = _slot_index(start_time, start_minutes, slot_minutes)
            end_slot = max(_slot_index(end_time, start_minutes, slot_minutes), start_slot + 1)
            start_col = 3 + start_slot
            end_col = min(2 + slot_count, 3 + end_slot - 1)
            if end_col < start_col:
                continue

            fill = PatternFill("solid", fgColor=CATEGORY_COLORS.get(item.get("category", ""), "F3E8D2"))
            for col in range(start_col, end_col + 1):
                sheet.cell(row=row, column=col).fill = fill
                sheet.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

            sheet.cell(row=row, column=start_col, value=item.get("label", ""))

        row += 1

    _set_widths(sheet, {1: 14, 2: 22})
    _wrap_all(sheet)


def _write_route_gantt(sheet, plan: dict) -> None:
    _title_cell(sheet, 1, "路线甘特图")
    start_minutes = 8 * 60
    end_minutes = 22 * 60
    slot_minutes = 30

    sheet.cell(row=3, column=1, value="日期")
    sheet.cell(row=3, column=2, value="主题")
    sheet.cell(row=3, column=1).font = Font(bold=True)
    sheet.cell(row=3, column=2).font = Font(bold=True)

    slot_count = int((end_minutes - start_minutes) / slot_minutes)
    for idx in range(slot_count):
        minute = start_minutes + idx * slot_minutes
        hour = minute // 60
        mins = minute % 60
        col = 3 + idx
        sheet.cell(row=3, column=col, value=f"{hour:02d}:{mins:02d}")
        sheet.cell(row=3, column=col).font = Font(bold=True, size=9)
        sheet.column_dimensions[get_column_letter(col)].width = 5

    row = 4
    for day in plan.get("daily_plan", []):
        sheet.cell(row=row, column=1, value=day.get("date", ""))
        sheet.cell(row=row, column=2, value=day.get("theme", ""))

        route_items = [
            item
            for item in day.get("items", [])
            if item.get("category") in {"交通", "缓冲"} or item.get("is_buffer")
        ]

        for item in route_items:
            start_time = item.get("start_time", "")
            end_time = item.get("end_time", "")
            if not start_time or not end_time:
                continue
            start_slot = _slot_index(start_time, start_minutes, slot_minutes)
            end_slot = max(_slot_index(end_time, start_minutes, slot_minutes), start_slot + 1)
            start_col = 3 + start_slot
            end_col = min(2 + slot_count, 3 + end_slot - 1)
            if end_col < start_col:
                continue

            fill = PatternFill("solid", fgColor=CATEGORY_COLORS.get(item.get("category", ""), "F3E8D2"))
            for col in range(start_col, end_col + 1):
                sheet.cell(row=row, column=col).fill = fill
                sheet.cell(row=row, column=col).alignment = Alignment(horizontal="center", vertical="center")

            sheet.cell(row=row, column=start_col, value=_route_label_for_gantt(item))

        row += 1

    _set_widths(sheet, {1: 14, 2: 22})
    _wrap_all(sheet)


def _slot_index(clock: str, start_minutes: int, slot_minutes: int) -> int:
    hour, minute = [int(part) for part in clock.split(":")]
    total = hour * 60 + minute
    return max(0, int((total - start_minutes) // slot_minutes))


def _route_label_for_gantt(item: dict) -> str:
    mode_label = str(item.get("route_mode_label") or item.get("route_mode") or "").strip()
    label = str(item.get("label", "") or "").strip()
    if item.get("category") == "交通" and mode_label:
        return f"{label} ({mode_label})"
    return label


def _display_route_provider(value) -> str:
    text = str(value or "").strip()
    mapping = {
        "amap": "高德地图",
        "google": "Google Maps",
        "manual": "保守时间预留",
    }
    return mapping.get(text.lower(), text)


def _route_distance_text(value) -> str:
    try:
        distance = float(value or 0)
    except (TypeError, ValueError):
        return ""
    if distance <= 0:
        return ""
    if distance >= 100:
        return f"{distance:.0f}"
    return f"{distance:.1f}"


def _set_widths(sheet, widths: dict[int, int]) -> None:
    for col, width in widths.items():
        sheet.column_dimensions[get_column_letter(col)].width = width


def _wrap_all(sheet) -> None:
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _stringify(value) -> str:
    if isinstance(value, list):
        return "、".join(str(item) for item in value if item)
    return str(value or "")
